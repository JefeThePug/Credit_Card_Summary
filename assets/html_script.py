import os
import pandas as pd
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

PY_PATH = os.getcwd()
APP_PATH = os.path.dirname(PY_PATH)

# Category Reference
charges = {('ＭＶ', 'マツクスバリユ', 'ライフ', 'スーパー', 'フレシコ', 'イズミヤ', '出原', 'サンディ', 'からあげ', 'ふくちぁん', 'フレスコ') :["食費", "食費"],
           ('マクドナルド', 'ＭＤ', 'すき家', 'ミスタードーナツ', 'やよい') :["外食費", "外食費"],
           ('飲料自販機', 'ＫＤＤＩ', 'ｃｈａｒｍ', 'モバイルＳｕｉｃａ', 'ジャパンファーマシー') :["Bさん用", "Bさん用"],
           ('薬局', 'ホームセンター', 'NITORI', 'カインズホーム', 'ニトリ', 'サンドラッグ') :["日用品", "日用品"],
           ('＃', 'ペテモ', '動物病院', 'ペット') :["愛犬用", "愛犬用"],
           ('ＡＰＰＬＥ', 'Ａｍａｚｏｎプライム') : ["サブスク", "サブスク"],
           ('保険',) :["Bさんの保険", "Bさんの"],
           ('ＢＩＧＬＯＢＥ利用料',) : ["インターネット", "インターネット"], 
           ('ａｕ携帯電',) :["Aさんの携帯端末代", "Aさんの携帯端末代"],
           ('イレブン', 'ローソン', 'ファミリーマート') :["コンビニ", "タバコ"],
           ('関西電力',) :["電気", "電気"],
           ('大阪ガス',) :["ガス", "ガス"],
           ('ＡＭＡＺＯＮ', 'ダイソー', 'ＥＴＣ', 'イデミツ') : ["その他", "その他"],
           ('駐車場', 'パーキング', 'タイムズ', 'パーク') : ["駐車代", "駐車代"]
          }

# Read HTML from file
with open(PY_PATH + '/html.txt', 'r') as f:
    contents = f.read()

# Make HTML Soup
soup = bs(contents, "lxml")
table = soup.find('table', id='meisaiTable')
headers = ["ご利用日", "ご利用店名", "ご利用金額", "カテゴリ"]
body = table.select('tbody tr')
name = soup.find("a",{"class":"default_btn btn_big_size co_btn size150 font_n mr_r40"}).get("href")
name = list(name.split("&s=")[1])
name.insert(-2, "-")
name = "".join(name)

trs = []
for i in body:
    j = i.getText()
    if "様" not in j:
        entry = [t for t in j.split('\n') if t]
        price = entry[2].replace(",","")
        if price.isnumeric():
            price = int(price)
        elif len(entry) >= 10:
            price = entry[9].replace(",","")
            if price.isnumeric():
                price = int(price)
            else:
                print(f"Error with the line \n{entry}")
                price = 0
        else:
            print(f"Error with the line \n{entry}")
            price = 0
        entry[2] = price
        trs.append(entry[:3])

while trs[-1][0] == "＜お支払金額総合計＞":
    trs = trs[:-1]

found = False
for item in trs:
    found = False
    for k, v in charges.items():
        for shop in k:
            condensed = "".join(c for c in shop if not c.isspace())
            if condensed in item[1]:
                item.append(v[bool(item[2]//5800)])
                found = True
                break
        if found: break
    if not found: item.append("")
    
trs.insert(0, ["","","",""])

df = pd.DataFrame(trs, columns=headers)

widths = (14, 49, 14, 14)

df = df[headers]

path = f"{APP_PATH}/ExcelPasterApp.xlsm"
with open(path):
    pass

wb = load_workbook(filename = path, keep_vba=True)
ws = wb.create_sheet(name)


for i in range(4):
    ws.column_dimensions[get_column_letter(i+1)].width = widths[i]
    ws.column_dimensions[get_column_letter(i+1)].font = Font(name="Meiryo")
    ws.column_dimensions[get_column_letter(i+1)].font = Font(size=14)

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
    
wb.save(path)